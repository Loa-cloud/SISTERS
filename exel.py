import spacy
import pandas as pd
import openpyxl
nlp = spacy.load('ru_core_news_sm')
from bs4 import BeautifulSoup
from pyquery import PyQuery as pq
from openpyxl import load_workbook
from openpyxl.styles import Font

import pickle
from tkinter import *
from tkinter import messagebox
from tkinter import scrolledtext
from tkinter import filedialog as fd
'''
import spacy
nlp = spacy.load('ru_core_news_sm')
from urllib.request import urlopen
from bs4 import BeautifulSoup
#from pyquery import PyQuery as pq
from openpyxl import load_workbook
from openpyxl.styles import Font
import pickle
from tkinter import *
from tkinter import messagebox
from tkinter import scrolledtext
from tkinter import filedialog as fd'''






filename = ""

# принимает в качестве аргумента файл, который нужно открыть
# функция открывает документ  возвращает его
def open_lem_cut_file(file):
    file2 = ""
    with open(file, 'r', encoding= "utf8")  as t:
        nums = t.read()
    file2 += nums
    lem_file = lem(file2)
    return (cut_dict(lem_file))


# принимает в качестве аргумента троку
# возвращает отсортированный словарь с лемматизацией
def lem(doc):
    doc = nlp(doc.lower())
    list1 =[token.lemma_ for token in doc]
    dic = {}
    for word in list1:
        if word.isalpha():
            if dic.get(word):
                dic[word] = dic[word] + 1
            else:
                dic[word] = 1
    dic = dict(sorted(dic.items(), key=lambda x: x[1]) )
    return(dic)

# принимает словарь
# возвращает урезанный словарь без 10% самых редко употребляемых слов
# и без 10% самых часто употребляемых слов
def cut_dict(dict_name):
    l = int(len(dict_name) * 0.1)
    len_hhk = len(dict_name)
    new_dict_name = {}
    i = 0
    for x in dict_name:
        i+=1
        if (i > l) and (i < (len_hhk-l)):
            new_dict_name[x] = dict_name[x]
    return new_dict_name

# принимает словарь по области и словарь по тексту
# возвращает коэфицент соответствия
def baes (dict_obl, dict_txt):
    count = 0
    for word in dict_txt:
        if word in dict_obl:
            count+=1
    try:
        q = count/len(dict_txt)
    except:
        q = count
    return q





def get_text_from_site(url):
    html = urlopen(url).read()
    soup = BeautifulSoup(html, features="html.parser")

    for script in soup(["script", "style"]):
        script.extract()    # rip it out

    text = soup.get_text()
    
    lines = (line.strip() for line in text.splitlines())
    chunks = (phrase.strip() for line in lines for phrase in line.split("  "))
    text = '\n'.join(chunk for chunk in chunks if chunk)
    
    return(text)



def load_dictionary(doc):   
    # загрузка словаря из файла
    with open(doc, 'rb') as file:
        myvar = pickle.load(file) 
    return myvar


def choose_exel():
    global filename
    filetypes = (('Эксель', '*.xlsx'),('Любой', '*'))

    filename = fd.askopenfilename(filetypes=filetypes)

def analiz():
    global filename
    wb = load_workbook(filename)
    #wb = load_workbook("1. Компании.xlsx")
    wb.sheetnames
    ws = wb.active

    incon_count = 0

    for i in range(100):#range(ws.max_row):
        try:
            site = ws.cell(row=i+2, column=8).value
            text = get_text_from_site(site)
        except:
            text = ws.cell(row=i+2, column=9).value
        #################################
        content = lem(str(text))


        d = {'КУЛЬТУРА И ИСКУССТВО': baes(art, content),
             'ЖКХ': baes(hhk, content),
             'БЕЗОПАСНОСТЬ И КОНТРОЛЬНО-НАДЗОРНАЯ ДЕЯТЕЛЬНОСТЬ': baes(safety, content),
             'ДОБЫЧА ПОЛЕЗНЫХ ИСКОПАЕМЫХ': baes(main, content),
             'ЗДРАВООХРАНЕНИЕ': baes(helth, content),
             'ИССЛЕДОВАНИЯ И ИНЖЕНЕРНО-ТЕХНИЧЕСКОЕ ПРОЕКТИРОВАНИЕ': baes(engineer, content),
             'КОНТЕНТ И СМИ': baes(news, content),
             'ОБРАЗОВАНИЕ': baes(study, content),
             'ПРОФЕССИОНАЛЬНАЯ ДЕЯТЕЛЬНОСТЬ ПРОЧАЯ': baes(other, content),
             'СЕЛЬСКОЕ ХОЗЯЙСТВО (ВКЛЮЧАЯ РЫБОЛОВСТВО)': baes(agricultural, content),
             'СПОРТ, ТУРИЗМ И СФЕРА ОБСЛУЖИВАНИЯ': baes(sport, content)}

        d = dict(sorted(d.items(), key=lambda x: -x[1]) )


        top3 = []
        top3 = list(d)[0:3]



        #################################    

        flag = False

        if ws.cell(row=i+2, column=3).value not in top3:
            ft = Font(color="FF0000")
            ws.cell(row=i+2, column=3).font = ft
            incon_count +=1

    wb.save("result1.xlsx")
    #print("Совпадений найдено:")
    #print(incon_count)

    wb_new = openpyxl.Workbook()
    ws = wb_new.active
    ws.title = "Статистика"
    ws["A1"] = "Количество найденных несоответствий"
    ws["B1"] = incon_count
    wb_new.save("statistika.xlsx")

''' def change_color(x):
    t = ""
    #ws = wb.active
    t = Font(color="FF0000")
    ws.cell(row=x, column=3).font = ft'''


    

Tk().withdraw()

art = load_dictionary('art.pkl')
hhk = load_dictionary('hhk.pkl')
safety = load_dictionary('safety.pkl')
main = load_dictionary('main.pkl')
helth = load_dictionary('helth.pkl')
engineer = load_dictionary('engineer.pkl')
news = load_dictionary('news.pkl')
study = load_dictionary('study.pkl')
other = load_dictionary('other.pkl')
agricultural = load_dictionary('agricultural.pkl')
sport = load_dictionary('sport.pkl')

window = Tk()
window.title('Хакатон')
window.geometry('300x200')
frame = Frame(
   window,
   padx=10,
   pady=10
)
frame.pack(expand=True)

cal_btn1 = Button(
frame,
text='Выбрать файл',
command = choose_exel
)
cal_btn1.grid(row=1, column=1)

cal_btn2 = Button(
frame,
text='Проанализировать',
command = analiz
)
cal_btn2.grid(row=2, column=1)

window.mainloop()








